"""
Это парсер данных с авито. Он проходится по каждому объявлению и собирает нужные
данные. Во время парсинга авито может указывать количество объявлений указанные
по параметрам с рекламой, так что иногда счётчик может показывать неправильное количество.
Парсер принимает параметры из xlsx файла. Его стандарт должен выглядить примерно так:
     Ключи     |     Регион    |     Категория     |     Подкатегория
---------------------------------------------------------------------------------------
 Дезинфекция  | Москва       | Услуги             | Уборка
 ...
"""
from urllib import parse
from sys import exit
from typing import List, NewType
from random import choice, random, randint
import json
import datetime
import time

import requests
from requests import Response
from bs4 import BeautifulSoup
from termcolor2 import c
import transliterate
import openpyxl

# CONSTANTS
HOST: str = 'https://www.avito.ru'
URL: str = ''
USER_AGENTS: List['user_agent'] = [user_agent.strip() for user_agent in open('user-agents.txt').readlines()]
MONTH_IN_NUMBER = {
    'января': '01',
    'февраля': '02',
    'марта': '03',
    'апреля': '04',
    'мая': '05',
    'июня': '06',
    'июля': '07',
    'августа': '08',
    'сентября': '09',
    'октября': '10',
    'ноября': '11',
    'декабря': '12'
}
# ANNOTATIONS
url_type = NewType('url_type', str)
int_price = NewType('int_price', int)
# To read the first fifty ads
counter_first_fifty_ads: int = 50
max_parsed_ad_on_one_session = 100
list_statistic_about_ad: List[dict] = []
start_time_script = time.time()
output_xlsx_file: dict = {}
parsed_ads: List[url_type] = []
names_columns_in_statistic_by_ad = [
    'Позиция (сортировка по умолчанию)',
    'Название объявления',
    'Описание',
    'Адрес',
    'Цена(в рублях)',
    'Имя продовца',
    'Количество объявлений у продовца',
    'Количество просмотров всего',
    'Количество просмотров за сегодня',
    'Ссылка на объявление',
    'Дата публикации',
    'Время публикации'
]


def get_response(url: url_type, params: dict = {}) -> Response or False:
    """Tries to get an answer by url from Avito 10 times. If it failed, it returns False"""
    number_of_request_try: int = 10
    for try_ in range(number_of_request_try):
        try:
            response: Response = requests.get(url, headers={'User-Agent': choice(USER_AGENTS)},
                                              timeout=60, params=params)
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
            print('Повторяем попытку...')
        except requests.exceptions.MissingSchema:
            print('Ссылка невалидна.')
        else:
            return response
    else:
        return False


def process_url_for_validity(ad_name: str, region: str, category: str,
                             subcategory: str, url: url_type) -> url_type or False:
    """
    Process url for parser. Avito generates a lot of noise when parsing data from its data.
    Sometimes the parameters by which the parser will search for data are invalid for the link,
    and this function does everything possible to create the correct link.

    ad_name: str -> ad name parameter.
    region: str -> region by which will search ads.
    category: str -> category by which will search ads.
    subcategory: str -> subcategory by which will search ads.

    @return: url or False -> If url success handled, return handled url else parameters incorrect, return False and
    stop parsing ads by this parameters.
    """
    for try_number in range(1, 4 + 1):
        try:
            test_request = get_response(url)
            bs_html: BeautifulSoup = BeautifulSoup(test_request.content, 'html.parser')
            test_get_text = bs_html.find('span', class_='page-title-count-1oJOc').get_text

        except AttributeError:
            if bs_html.find('h2', class_='no-results-title-3kn6E') is not None:

                print('Ничего не найдено в выбранной области поиска')
                return False

            elif bs_html.find('h2', class_='firewall-title') is None:

                if try_number == 1:

                    past_region = region
                    region = region.split()[-1].strip()
                    region_for_url = transliterate.translit(
                        region, reversed=True).replace(' ', '_').strip().replace("'", '').replace('j', 'y')
                    url = f"{HOST}/{region_for_url}{url[21 + url[21:].find('/'):]}"

                elif try_number == 2:
                    print(f'Введите регион ({region}), что бы он отвечал на вопрос (область чья?) '
                          'или образовывал словосочитание по типу (Московская область).\n'
                          'Примеры:\n'
                          ' Киров -> Кировская\n'
                          ' Москва -> Московская\n')
                    region = input('Вводите: ').strip() + ' область ' + region
                    region_for_url = transliterate.translit(
                        region, reversed=True).replace(' ', '_').strip().replace("'", '').replace('j', 'y')

                    url = f"{HOST}/{region_for_url}{URL[21 + URL[21:].find('/'):]}"

                elif try_number == 3:

                    print('Введите ссылку вручную. Её можно получить на Авито '
                          'https://www.avito.ru/ в адресной строке, указав в поисковик параметры:\n'
                          f'Ключ: {ad_name}; Регион: {past_region}; Категория: {category}; Подкатегория: {subcategory};\n')
                    url = input('Вводите: ')
                    region = transliterate.translit(URL[21:URL.find('/', 21)], 'ru')

                else:
                    print(c('По вашим параметрам ничего не найдено.').red)
                    return False
            elif bs_html.find('h2', class_='firewall-title') is not None:
                print(c('Ваш IP адрес заблокировал Avito на время. Следуйте указаниям файла help.txt.').red)
                exit()

        else:
            return url


def generate_valid_url_for_parsing(ad_name: str, region: str, category: str, subcategory: str) -> url_type:
    """Generate url for parsing ads from avito."""
    global URL

    # Parameters may be with comma. Them need delete.
    if region.count(',') == 1:
        region = region.replace(',', '')
    elif region.count(',') > 1:
        region = region.split(',')[-1].strip()

    ad_name_for_url = parse.urlencode({'q': ad_name})
    region_for_url = transliterate.translit(
        region, reversed=True).replace(' ', '_').strip().replace("'", '').replace('j', 'y')

    # Avito for this two category make unique name in link
    if category == 'услуги':
        category_for_url = 'predlozheniya_uslug'
    elif category == 'готовый бизнес и оборудование':
        category_for_url = 'dlya_biznesa'
    else:
        category_for_url = transliterate.translit(
            category, reversed=True).replace(' ', '_').strip().replace("'", '').replace('j', 'y')

    try:
        if subcategory == 'уборка':
            subcategory_for_url = 'uborka_klining'
        else:
            subcategory_for_url = transliterate.translit(
                subcategory, reversed=True).strip().replace(' ', '_').replace("'", '').replace('j', 'y')
    except transliterate.exceptions.LanguageDetectionError:
        subcategory_for_url = ''
    handled_url = process_url_for_validity(
        ad_name, region, category, subcategory,
        f'{HOST}/{region_for_url}/{category_for_url}/{subcategory_for_url}?{ad_name_for_url}')
    URL = handled_url


def set_common_amount_of_ad() -> None:
    """Set common amount of ad."""

    page_with_ads = get_response(URL)
    bs_html: BeautifulSoup = BeautifulSoup(page_with_ads.content, 'html.parser')
    amount_ad = bs_html.find('span', class_='page-title-count-1oJOc')
    output_xlsx_file['Общее количество объявлений'] = int(amount_ad.string.replace(' ', ''))


def set_total_amount_views(views_on_ad_page: list) -> None:
    """
    Set total amount views on today and in generally first 50 ads and all.

    views_on_ad_page: list -> list consist of views all time and today. ['{views all time}', '(+{views today})']
    """
    global counter_first_fifty_ads

    if counter_first_fifty_ads != 0:
        output_xlsx_file['Общее количество просмотров первых 50 объявлений (всего)'] += int(
            views_on_ad_page[0])
        output_xlsx_file['Общее количество просмотров первых 50 объявлений (сегодня)'] += int(
            views_on_ad_page[1][2:-1])
        counter_first_fifty_ads -= 1
    output_xlsx_file['Общее количество просмотров всего'] += int(views_on_ad_page[0])
    output_xlsx_file['Общее количество просмотров за сегодня'] += int(views_on_ad_page[1][2:-1])


def add_price_to_price_from_all_ads(price_of_product: int_price) -> None:
    """
    Adds the price to the price from all ads.

    price_of_product: int_price -> price specified in the ad.
    """
    output_xlsx_file['Средняя цена всех со всех объявлений'] += price_of_product


def set_date_of_publication_of_ad() -> None:
    """Sets the date of publication of the ad at 10, 20 and 50 places."""
    list_of_ad_publication_dates = []
    max_number_of_attempts = 10

    while not list_of_ad_publication_dates:
        time.sleep(3)

        ads_sorted_by_date = get_response(URL, {'s': 104})
        bs_page_ads = BeautifulSoup(ads_sorted_by_date.content, 'html.parser')
        ads_publication_dates = bs_page_ads.find_all('div', class_='snippet-date-info')

        if max_number_of_attempts == 0:
            print('Повторите попытку.')
            exit()

        max_number_of_attempts -= 1

        for date_tag in ads_publication_dates[:output_xlsx_file['Общее количество объявлений']]:
            try:
                day_publication, month_publication = date_tag['data-tooltip'].split()[:2]
            except ValueError:
                continue
            month_publication = MONTH_IN_NUMBER[month_publication]
            list_of_ad_publication_dates.append(f'{day_publication}.{month_publication}.2020')

    # If there are fewer ads, the script tries to install what it can.
    try:
        output_xlsx_file['Дата публикации 10-ого объявления (сортировка по дате)'] = list_of_ad_publication_dates[10]
    except IndexError:
        return

    try:
        output_xlsx_file['20-ого объявления (сортировка по дате)'] = list_of_ad_publication_dates[20]
    except IndexError:
        return

    try:
        output_xlsx_file['50-ого объявления (сортировка по дате)'] = list_of_ad_publication_dates[-1]
    except IndexError:
        return


def set_average_price_of_all_ads() -> None:
    """Sets and calculates the arithmetic average of prices from all ads."""
    output_xlsx_file['Средняя цена всех со всех объявлений'] = \
        (output_xlsx_file['Средняя цена всех со всех объявлений']
         / output_xlsx_file['Общее количество объявлений'])


def date_handler(date_publication: str) -> str:
    """Handled received date from ad. Return data in format day.month.year."""
    datetime_today = datetime.date.today()
    datetime_for_format = datetime.date(datetime_today.year, datetime_today.month, datetime_today.day)

    if date_publication.lower() == 'сегодня':
        date_publication = datetime_for_format.strftime('%d.%m.%Y')
    elif date_publication.lower() == 'вчера':
        date_publication = str(datetime_today.day - 1) + datetime_for_format.strftime('.%m.%Y')
    else:
        day_publication, month_publication = date_publication.split()
        month_publication = MONTH_IN_NUMBER[month_publication]
        date_publication = f'{day_publication}.{month_publication}.2020'
    return date_publication


def set_data_about_ad(ad_page: Response, views_on_ad: list, price_of_product: int_price, link: url_type) -> None:
    """Fill dictionary need data and add filled dictionary in list."""
    bs_ad_page: BeautifulSoup = BeautifulSoup(ad_page.content, 'html.parser')

    try:
        total_views: int = int(views_on_ad[0])
        views_today: int = int(views_on_ad[1][2:-1])
    except IndexError:
        total_views = 0
        views_today = 0

    statistic_about_ad = {
        'Название объявления': '',
        'Описание': '',
        'Адрес': '',
        'Цена': price_of_product,
        'Имя продовца': '',
        'Количество объявлений у продовца': 0,
        'Количество просмотров всего': total_views,
        'Количество просмотров за сегодня': views_today,
        'Ссылка на объявление': link,
        'Дата публикации': '',
        'Время публикации': ''
    }

    try:
        ad_description: str = bs_ad_page.find('div', class_='item-description').get_text().strip()
    except AttributeError:
        ad_description: str = 'Нету'
    try:
        seller_address: str = bs_ad_page.find('div', class_='item-address').get_text().strip()
    except AttributeError:
        seller_address: str = 'Нету'
    try:
        button_with_number_active_ads = bs_ad_page.find(class_='seller-info-favorite-seller-buttons').get('data-props')
        number_active_ads = int(json.loads(button_with_number_active_ads)['summary'].split()[0])
    except (AttributeError, IndexError):
        number_active_ads = 0

    ad_title: str = bs_ad_page.find('span', class_='title-info-title-text').string.strip()
    date_and_time_publication = bs_ad_page.find('div', class_='title-info-metadata-item-redesign').string.strip()
    seller_name: str = bs_ad_page.find('div', class_='seller-info-name').get_text()
    publication_date = date_handler(date_and_time_publication[:-5].strip().replace(' в', ''))
    publication_time = date_and_time_publication[-5:]  # 16:24

    statistic_about_ad['Название объявления'] = ad_title
    statistic_about_ad['Описание'] = ad_description
    statistic_about_ad['Адрес'] = seller_address
    statistic_about_ad['Имя продовца'] = seller_name
    statistic_about_ad['Количество объявлений у продовца'] = number_active_ads
    statistic_about_ad['Дата публикации'] = publication_date
    statistic_about_ad['Время публикации'] = publication_time

    list_statistic_about_ad.append(statistic_about_ad)


def bypass_traps_avito(bs_ad_html: BeautifulSoup, ad_page: Response, link: url_type) -> None:
    """
    Bypasses the traps that Avito makes. For example, it can send that
     the page was not found, although it exists. If the price of the ad
     does not indicate the product price is 0. If Avito sends an error 429,
     the script stops working because it can no longer send requests due to IP blocking.

    bs_ad_html: BeautifulSoup -> object from BeautifulSoup from ad page.
    ad_page: Response -> response from ad page Avito.
    link: str -> link on ad page.
    """
    try:
        views_on_ad_page: str = bs_ad_html.find(class_='title-info-metadata-item').get_text()[1:].split()
        try:
            price_of_product: int_price = int(
                str(bs_ad_html.find('span', class_='js-item-price').text).replace(' ', ''))
        except AttributeError:
            price_of_product: int_price = 0
    except AttributeError:
        if ad_page.status_code == 404:
            try:
                ad_page: Response = get_response(link)
                bs_ad_html: BeautifulSoup = BeautifulSoup(ad_page.content, 'html.parser')
                views_on_ad_page: list = bs_ad_html.find(class_='title-info-metadata-item').get_text()[1:].split()
                try:
                    price_of_product: int_price = int(
                        str(bs_ad_html.find('span', class_='js-item-price').text).replace(' ', ''))
                except AttributeError:
                    price_of_product: int_price = 0
            except AttributeError:
                print(c('''
                      Это исключение возможно лишь при том, если Avito отвечает не тем сайтом, который пар-
                      сер ожидал увидеть. За остальной информацией обращайтесь к файлу help.txt.

                      Если программа продолжает парсить, необращайте внимания.
                      ''').yellow)
            else:
                try:
                    set_total_amount_views(views_on_ad_page)
                except IndexError:
                    pass
                add_price_to_price_from_all_ads(price_of_product)
                set_data_about_ad(ad_page, views_on_ad_page, price_of_product, link)
        elif ad_page.status_code == 429:
            print(c('''
                  Ваш IP был заблокирован на время. Нужно подождать некоторое время, либо зайти на
                  на сайт через браузер и ввести капчу. Если ничто из этого не помогло, следуйте указа-
                  ниям в файле help.txt.
                  ''').red)
            exit()
    else:
        try:
            set_total_amount_views(views_on_ad_page)
        except IndexError:
            pass
        add_price_to_price_from_all_ads(price_of_product)
        set_data_about_ad(ad_page, views_on_ad_page, price_of_product, link)


def send_ad_data_to_functions(max_pages: int) -> None:
    """
    Send data to functions, received from ad page. In this function implementation
     simulation of the human factor (makes random pauses before the request
     (about 8-10 seconds) and requests a random number of times (from 1 to 5)).
     This is done so that Avito does not consider the parser a bot. Also protects
     against advertisements from another city that spoil statistics. He function
     goes through the pages (if the number of ads allows), opens the ads and sends
     the desired data to the functions. If Avito give sponsored links, function skip this.

     max_pages: int -> maximum number of pages with ads.
    """
    maximum_amount_of_open_links_without_pause: int = randint(1, 5)
    next_page: int = 1
    counter_parsed_link: int = 1

    # Sometimes ads are not enough for one page and Avito fills it with similar ads.
    if output_xlsx_file['Общее количество объявлений'] <= 50:
        ad_limit = output_xlsx_file['Общее количество объявлений']
    else:
        ad_limit = 'More than 50'

    while next_page != max_pages + 1:

        link_on_page_with_ads: url_type = get_response(URL).url

        # Avito does not allow you to navigate through the pages if a link with random letters is not specified
        # (this is how it looks like avito.ru/chita/krasota_i_zdorove/kupit-meditsinskie_izdeliya-ASgBAgICAUSEAqgJ),
        # this link can be obtained by request. This is the only way to go through the pages
        if '?' in link_on_page_with_ads:
            link_to_navigate_through_pages: url_type = link_on_page_with_ads + f'&p={next_page}'
        else:
            link_to_navigate_through_pages: url_type = link_on_page_with_ads + f'?p={next_page}'
        time.sleep(2.2)

        page_with_ad: Response = get_response(link_to_navigate_through_pages)
        content_of_ad_page: BeautifulSoup = BeautifulSoup(page_with_ad.content, 'html.parser')
        links_on_ads: List[url_type] = []

        for element in content_of_ad_page.find_all(class_='item_table-description'):
            link_on_ad = HOST + element.find('a').get('href')
            if link_on_ad[21:].find('/') == -1:
                continue
            else:
                links_on_ads.append(link_on_ad)

        if not links_on_ads:
            # Avito sometimes gives out the wrong data that it usually gives out (we are talking about classes).
            for element in content_of_ad_page.find_all(class_='iva-item-body-NPl6W'):
                link_on_ad = HOST + element.find('a').get('href')
                if link_on_ad[21:].find('/') == -1:
                    continue
                else:
                    links_on_ads.append(link_on_ad)
        if not links_on_ads:
            open('new_tags_неудалять.html', 'w').write(page_with_ad.text)
            print('Авито прислал невалидный сайт, перезапустите программу с теми параметрами, '
                  'на которых остановился парсер.')
            exit()

        time.sleep(3)

        for link in links_on_ads:
            if link not in parsed_ads:
                # It is done to imitate a person, so that Avito does not consider the parser a bot.
                # If delete this code, Avito can give block by IP for a while.
                if maximum_amount_of_open_links_without_pause == 0:
                    time.sleep(round(randint(8, 10) + random(), 2))
                    maximum_amount_of_open_links_without_pause = randint(1, 5)
                maximum_amount_of_open_links_without_pause -= 1

                ad_page: Response = get_response(link)
                bs_ad_content: BeautifulSoup = BeautifulSoup(ad_page.content, 'html.parser')
                bypass_traps_avito(bs_ad_content, ad_page, link)
                parsed_ads.append(link)

                print(c(f'{link[12:]: <115} спарсено удачно.').magenta,
                      f'Осталось {counter_parsed_link}/{output_xlsx_file["Общее количество объявлений"]}')
                counter_parsed_link += 1
                if isinstance(ad_limit, int):
                    ad_limit -= 1
                if ad_limit == 0:
                    break

        if isinstance(ad_limit, int): break
        print(f'{next_page} из {max_pages} спарсено.')
        next_page += 1

    set_average_price_of_all_ads()
    print(c('Парсинг завершен успешно!').green)


# Write to xlsx
def write_first_list(workbook_list) -> None:
    """Writes in first leaf (Общая статистика) received data in the process of parsing."""
    row: int = 2
    column_letters = 'ABCDEFGHIJK'
    row_is_free: bool = False

    while True:

        for col_letter in column_letters:
            if workbook_list[col_letter + str(row)].value is None:
                row_is_free = True
            else:
                row_is_free = False
                break

        if row_is_free:
            for column in range(1, 11 + 1):
                workbook_list.cell(row=row, column=column).value = list(output_xlsx_file.values())[column - 1]
            break
        else:
            row += 1


def write_second_list(workbook_list) -> None:
    """Writes in second leaf (Статистика по объявлениям) received data in the process of parsing."""
    row: int = 2
    column_letters = 'ABCDEFGHIJKL'
    row_is_free: bool = False
    counter_rows: int = 0
    
    while True:

        for col_letter in column_letters:
            if workbook_list[col_letter + str(row)].value is None:
                row_is_free = True
            else:
                row_is_free = False
                break
        if row_is_free:
            for statistic_ad in list_statistic_about_ad:
                counter_rows += 1
                workbook_list.cell(row=row, column=1).value = counter_rows
                for column in range(2, len(names_columns_in_statistic_by_ad) + 1):
                    workbook_list.cell(row=row, column=column).value = list(statistic_ad.values())[column - 2]
                row += 1
            break
        else:
            row += 1


def send_workbook_lists() -> None:
    """
    Send functions workbook list in which need write data. If the excel file
    table does not exist, it creates a new one with the required parameters and sheets.
    """
    xlsx_file_name: str = 'avito_statistic.xlsx'

    try:
        workbook = openpyxl.load_workbook(xlsx_file_name)
    except FileNotFoundError:
        new_workbook = openpyxl.Workbook()

        for sheet_name in new_workbook.sheetnames:
            sheet = new_workbook[sheet_name]
            new_workbook.remove(sheet)

        new_workbook_list_1 = new_workbook.create_sheet('Общая статистика')

        for column in range(1, 11 + 1):
            new_workbook_list_1.cell(row=1, column=column).value = list(output_xlsx_file.keys())[column - 1]

        new_workbook_list_2 = new_workbook.create_sheet('Статистика по объявлениям')

        for column in range(1, len(names_columns_in_statistic_by_ad) + 1):
            new_workbook_list_2.cell(row=1, column=column).value = names_columns_in_statistic_by_ad[column - 1]

        new_workbook.save(xlsx_file_name)
        new_workbook.close()
        workbook = openpyxl.load_workbook(xlsx_file_name)

    workbook_list_1 = workbook['Общая статистика']
    workbook_list_2 = workbook['Статистика по объявлениям']

    write_first_list(workbook_list_1)
    write_second_list(workbook_list_2)

    workbook.save(xlsx_file_name)
    workbook.close()


def run():
    global list_statistic_about_ad, counter_first_fifty_ads, output_xlsx_file, parsed_ads, URL

    row: int = 2
    column_letters = 'ABCD'
    row_is_fill: bool = False

    xlsx_file_path = input('Введите название файла с параметрами или его путь: ')
    workbook = openpyxl.load_workbook(xlsx_file_path)
    workbook_list = workbook['Лист1']

    while True:

        for col_letter in column_letters:
            if workbook_list[col_letter + str(row)].value is None:
                row_is_fill = False
            else:
                row_is_fill = True
                break

        if row_is_fill:
            print('Подождите 40-65 секунд...')
            time.sleep(round(randint(40, 65) + random(), 2))

            # Parameters for link which need parsed
            ad_name = workbook_list[f'A{row}'].value
            region = workbook_list[f'B{row}'].value.lower().strip()
            category = workbook_list[f'C{row}'].value.lower().strip()
            subcategory = workbook_list[f'D{row}'].value.lower().strip()

            generate_valid_url_for_parsing(ad_name, region, category, subcategory)
            if not URL:
                continue

            row += 1

            # Reset past data
            list_statistic_about_ad = []
            counter_first_fifty_ads = 50
            parsed_ads = []
            output_xlsx_file = {
                'Ключ': ad_name,
                'Регион': region,
                'Общее количество объявлений': 0,
                'Общее количество просмотров всего': 0,
                'Общее количество просмотров за сегодня': 0,
                'Дата публикации 10-ого объявления (сортировка по дате)': 'Нету',
                '20-ого объявления (сортировка по дате)': 'Нету',
                '50-ого объявления (сортировка по дате)': 'Нету',
                'Средняя цена всех со всех объявлений': 0,
                'Общее количество просмотров первых 50 объявлений (сегодня)': 0,
                'Общее количество просмотров первых 50 объявлений (всего)': 0
            }

            print(f'Парсится ссылка {URL}')
            print(f'{"Ключ": ^25} | {"Регион": ^26} | {"Категория": ^29} | {"Подкатегория": ^32}')
            print('_' * 121)
            print(f'{ad_name: ^23} | {region: ^26} | {category: ^28} | {subcategory: ^32}')

            set_common_amount_of_ad()
            set_date_of_publication_of_ad()
            time.sleep(3)

            avito_response: Response = get_response(URL)
            avito_page_content: BeautifulSoup = BeautifulSoup(avito_response.content, 'html.parser')
            try:
                max_pages = int(avito_page_content.find_all('span', class_='pagination-item-1WyVp')[-2].text)
            except IndexError:
                max_pages = 1

            send_ad_data_to_functions(max_pages)

            print('Записываем данные...')
            send_workbook_lists()
            print(c('Данные успешно записаны!').green)
        else:
            break

    print('Время работы парсера', round(time.time() - start_time_script), 'секунды')


if __name__ == '__main__':
    run()
